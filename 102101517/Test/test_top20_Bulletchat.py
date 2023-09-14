import openpyxl

# 打开并读取文本文件
with open('bulletchat.txt', 'r', encoding='utf-8') as file:
    lines = file.readlines()

# 创建一个默认字典来统计句子的出现次数
sentence_counts = {}

# 逐行遍历文本文件
for line in lines:
    # 去除行末尾换行符和空格
    sentence = line.strip()
    # 排除只包含数字的行
    if not sentence.isdigit():
        # 将每行的内容作为字典的键，递增相应的值
        if sentence in sentence_counts:
            sentence_counts[sentence] += 1
        else:
            sentence_counts[sentence] = 1

# 对字典按值进行降序排序
sorted_sentence_counts = sorted(sentence_counts.items(), key=lambda x: x[1], reverse=True)
# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Bulletchat Counts"
# 将每行文本和其出现次数写入Excel文件
for row_number, (line, count) in enumerate(sorted_sentence_counts, start=1):
    worksheet.cell(row=row_number, column=1, value=line)
    worksheet.cell(row=row_number, column=2, value=count)
# 保存Excel文件
workbook.save('Bulletchat_counts.xlsx')

# 输出出现次数前20的弹幕
top20_bulletchat = sorted_sentence_counts[:20]
for line, count in top20_bulletchat:
    print(f'{line}: {count}')
# 将前20行文字内容写入txt文本
with open('top20_Bulletchat.txt', 'w', encoding='utf-8') as txt_file:
    for line, _ in top20_bulletchat:
        txt_file.write(line + '\n')








